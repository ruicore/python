# -*- coding: utf-8 -*-
# @Date:   2018-05-25 9:24:50
# @Last Modified by:   何睿
# @Last Modified time: 2018-05-25 9:24:50

import urllib
import requests
import json
import openpyxl
import os
import re
import glob


def get_access_token():
    # client_id 为官网获取的AK， client_secret 为官网获取的SK
    host = 'https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=[Your key]&client_secret=[Your secret key]'
    request = urllib.request.Request(host)
    request.add_header('Content-Type', 'application/json; charset=UTF-8')
    response = urllib.request.urlopen(request)
    content = response.read()
    content = json.loads(content)
    if content:
        return content["access_token"]


'''接口接入，返回json格式数据'''


def get_content(text):
    access_token = get_access_token().strip()
    url = "https://aip.baidubce.com/rpc/2.0/nlp/v2/comment_tag?access_token=" + \
        access_token  # API
    headers = {"Content-Type": "application/json"}
    data = {"text": text, "type": 1}
    try:
        data = json.dumps(data)
        r = requests.post(url, data=data, headers=headers)
        return (r.text)
    except Exception as e:
        print(e)
        return 0


def get_all_files(directory, filetype=None):
    directorys = []
    for dir_path, dir_names, file_names in os.walk(directory):
        # 如果输入了文件类型，则按类型查找
        if filetype:
            try:
                directorys += glob.glob(dir_path + '\\*.' + filetype)
            except Exception as e:
                print(e)
        # 如果没有输入，则返回所有的文件
        else:
            for file_name in file_names:
                path = dir_path + "\\" + file_name
                directorys.append(path)
    return directorys


if __name__ == "__main__":
    files = get_all_files(r'C:\HeRui\Temp\data')
    replace = re.compile("[^0-9a-zA-Z\u4e00-\u9fa5.，,。？“”]")
    i = 1
    for file in files:
        excel = openpyxl.load_workbook(file)
        sheets = excel.sheetnames
        work_sheet = excel[sheets[0]]
        first_column = work_sheet['F']
        lines = []
        for cell in first_column[1:]:
            line=[]
            if cell.value:
                try:
                    baidu=get_content(replace.sub("", cell.value))
                    t = json.loads(baidu)
                    try:
                        for item in t['items']:
                            line.append(item["abstract"].split('>')[1].split('<')[0])
                    except Exception as e:
                        print(e)
                except Exception as e:
                    print(e) 
            else:
                pass
            lines.append(line)
        print(lines)
        new_excel=openpyxl.Workbook()
        new_sheet = new_excel.create_sheet(index=0)
        for item in lines:
            new_sheet.append(item)
        new_excel.save('C:\\HeRui\\Temp\\temp\\'+file.split('\\')[-1])
        i+=1
